import openpyxl

Dict = {}
book1 = openpyxl.load_workbook("C:\\Users\\ADMIN\\Desktop\\PythonDemo.xlsx")
sheet1 = book1.active
cell1 = sheet1.cell(row=1, column=2)
# print(cell1.value)
val = sheet1.cell(row=2, column=2).value = "Rahul"
# print(val)

rows = sheet1.max_row
# print(rows - 1)

columns = sheet1.max_column
# print(columns - 1)

cells = sheet1['A3'].value
# print(cells)

for i in range(1, sheet1.max_row+1):
    # if sheet1.cell(row=i, column=1).value == "testcase2":
    for j in range(2, sheet1.max_column+1):
        Dict[sheet1.cell(row=1, column=j).value] = sheet1.cell(row=i, column=j).value
print(Dict)
