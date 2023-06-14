import openpyxl


class HomePageData:

    test_homepage_data = [{"firstname": "Rahul", "email": "rahul@gmail.com", "gender": "Female"},
                          {"firstname": "Anshika", "email": "anshika@gmail.com", "gender": "Male"}]

    @staticmethod
    def gettestdata(test_case_name):
        dicti = {}
        book = openpyxl.load_workbook("C:\\Users\\ADMIN\\Desktop\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dicti[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dicti]
